"""
Tests for SFM graph exporters.

Tests XLSX export functionality for Hayden-compliant delivery matrices.
"""

import uuid
from pathlib import Path
import tempfile

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from api.sfm_service import SFMService
from models import Node
from models.delivery_matrix import Delivery, SFMDeliveryMatrix
from graph.exporters import export_delivery_matrix_to_xlsx


class TestXLSXExporter:
    """Test XLSX export for delivery matrices."""

    def setup_method(self):
        """Setup test service and sample delivery matrix."""
        # Create service
        self.service = SFMService()

        # Create sample components
        self.legislature = Node(label="Legislature", description="State legislative body")
        self.dept_ed = Node(label="Department of Education", description="State education agency")
        self.school_district = Node(label="School District", description="Local school district")

        self.service.create_node(self.legislature)
        self.service.create_node(self.dept_ed)
        self.service.create_node(self.school_district)

        # Create delivery matrix
        self.matrix = self.service.create_delivery_matrix(
            components=[self.legislature.id, self.dept_ed.id, self.school_district.id],
            description="K-12 Education Finance Matrix",
            label="Test Education Matrix"
        )

        # Add deliveries: Legislature -> School District
        money_delivery = Delivery(
            delivery_type="money",
            delivery_content="$800M annual appropriation via TEEOSA formula",
            quantity=800_000_000,
            units="USD/year",
            temporal_rate="annual",
            certainty=0.95
        )
        self.service.add_delivery_to_matrix(
            self.matrix,
            self.legislature.id,
            self.school_district.id,
            money_delivery,
            cell_description="Legislature provides funding to school districts through state aid formula"
        )

        rule_delivery = Delivery(
            delivery_type="rule",
            delivery_content="TEEOSA formula compliance requirements",
            certainty=1.0
        )
        self.service.add_delivery_to_matrix(
            self.matrix,
            self.legislature.id,
            self.school_district.id,
            rule_delivery,
            cell_description="Legislature provides funding to school districts through state aid formula"
        )

        # Add deliveries: Legislature -> Department of Education
        authority_delivery = Delivery(
            delivery_type="authority",
            delivery_content="Oversight and audit authority",
            certainty=1.0
        )
        self.service.add_delivery_to_matrix(
            self.matrix,
            self.legislature.id,
            self.dept_ed.id,
            authority_delivery,
            cell_description="Legislature grants oversight authority to Department of Education"
        )

        # Add deliveries: Department of Education -> School District
        info_delivery = Delivery(
            delivery_type="information",
            delivery_content="Academic performance standards and reporting requirements",
            temporal_rate="continuous"
        )
        self.service.add_delivery_to_matrix(
            self.matrix,
            self.dept_ed.id,
            self.school_district.id,
            info_delivery,
            cell_description="Department of Education sets standards for school districts"
        )

    def test_export_creates_file(self):
        """Verify XLSX file is created at specified path."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            assert tmp_path.exists()
            assert tmp_path.stat().st_size > 0

        finally:
            tmp_path.unlink()

    def test_export_has_three_sheets(self):
        """Verify workbook contains Matrix View, Cell Descriptions, Delivery Details."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames

            assert len(sheet_names) == 3
            assert "Matrix View" in sheet_names
            assert "Cell Descriptions" in sheet_names
            assert "Delivery Details" in sheet_names

        finally:
            tmp_path.unlink()

    def test_matrix_view_structure(self):
        """Verify N×N structure with correct headers."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            ws = wb["Matrix View"]

            # Check header cell
            assert ws.cell(1, 1).value == "Source \\ Target"

            # Check column headers (row 1)
            assert ws.cell(1, 2).value == "Legislature"
            assert ws.cell(1, 3).value == "Department of Education"
            assert ws.cell(1, 4).value == "School District"

            # Check row headers (column A)
            assert ws.cell(2, 1).value == "Legislature"
            assert ws.cell(3, 1).value == "Department of Education"
            assert ws.cell(4, 1).value == "School District"

        finally:
            tmp_path.unlink()

    def test_matrix_view_content(self):
        """Verify cell content shows delivery summaries."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            ws = wb["Matrix View"]

            # Check Legislature -> School District cell (row 2, col 4)
            cell_content = ws.cell(2, 4).value
            assert cell_content is not None
            assert "money:" in cell_content.lower()
            assert "$800M" in cell_content or "800M" in cell_content
            assert "rule:" in cell_content.lower()

            # Check Legislature -> Department of Education cell (row 2, col 3)
            cell_content = ws.cell(2, 3).value
            assert cell_content is not None
            assert "authority:" in cell_content.lower()

            # Check Department of Education -> School District cell (row 3, col 4)
            cell_content = ws.cell(3, 4).value
            assert cell_content is not None
            assert "information:" in cell_content.lower()

            # Check empty cell (School District -> Legislature)
            cell_content = ws.cell(4, 2).value
            assert cell_content == "" or cell_content is None

        finally:
            tmp_path.unlink()

    def test_matrix_view_color_coding(self):
        """Verify cells are colored by dominant delivery type."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            ws = wb["Matrix View"]

            # Legislature -> School District has 2 deliveries: money (dominant) and rule
            # Should be colored light green (money color)
            cell = ws.cell(2, 4)
            if cell.fill and cell.fill.start_color:
                fill_color = cell.fill.start_color.rgb or cell.fill.start_color.index
                # Light green is C6EFCE (money color) - openpyxl may add prefix
                if isinstance(fill_color, str):
                    assert "C6EFCE" in fill_color

            # Legislature -> Department of Education has 1 delivery: authority
            # Should be colored light yellow (authority color)
            cell = ws.cell(2, 3)
            if cell.fill and cell.fill.start_color:
                fill_color = cell.fill.start_color.rgb or cell.fill.start_color.index
                # Light yellow is FFEB9C (authority color) - openpyxl may add prefix
                if isinstance(fill_color, str):
                    assert "FFEB9C" in fill_color

        finally:
            tmp_path.unlink()

    def test_cell_descriptions_sheet(self):
        """Verify Cell Descriptions sheet has all non-empty cells."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            ws = wb["Cell Descriptions"]

            # Check headers
            assert ws.cell(1, 1).value == "Source"
            assert ws.cell(1, 2).value == "Target"
            assert ws.cell(1, 3).value == "Description"
            assert ws.cell(1, 4).value == "Delivery Count"

            # Should have 3 non-empty cells (Legislature->School, Legislature->Dept, Dept->School)
            # Each gets a row starting from row 2
            row_count = ws.max_row
            assert row_count >= 4  # Header + 3 data rows

            # Check Legislature -> School District row
            found_leg_school = False
            for row_idx in range(2, row_count + 1):
                source = ws.cell(row_idx, 1).value
                target = ws.cell(row_idx, 2).value
                if source == "Legislature" and target == "School District":
                    found_leg_school = True
                    desc = ws.cell(row_idx, 3).value
                    count = ws.cell(row_idx, 4).value
                    assert "funding" in desc.lower()
                    assert count == 2  # money + rule deliveries
                    break

            assert found_leg_school

        finally:
            tmp_path.unlink()

    def test_delivery_details_sheet(self):
        """Verify Delivery Details sheet has all deliveries."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)
            ws = wb["Delivery Details"]

            # Check headers
            assert ws.cell(1, 1).value == "Source"
            assert ws.cell(1, 2).value == "Target"
            assert ws.cell(1, 3).value == "Delivery Type"
            assert ws.cell(1, 4).value == "Content"
            assert ws.cell(1, 5).value == "Quantity"
            assert ws.cell(1, 6).value == "Units"
            assert ws.cell(1, 7).value == "Temporal Rate"
            assert ws.cell(1, 8).value == "Threshold"
            assert ws.cell(1, 9).value == "Certainty"

            # Should have 4 deliveries total (2 in Leg->School, 1 in Leg->Dept, 1 in Dept->School)
            row_count = ws.max_row
            assert row_count >= 5  # Header + 4 data rows

            # Check money delivery exists
            found_money = False
            for row_idx in range(2, row_count + 1):
                delivery_type = ws.cell(row_idx, 3).value
                if delivery_type == "money":
                    found_money = True
                    content = ws.cell(row_idx, 4).value
                    quantity = ws.cell(row_idx, 5).value
                    units = ws.cell(row_idx, 6).value
                    rate = ws.cell(row_idx, 7).value
                    certainty = ws.cell(row_idx, 9).value

                    assert "$800M" in content or "800M" in content
                    assert quantity == 800_000_000
                    assert units == "USD/year"
                    assert rate == "annual"
                    assert certainty == 0.95
                    break

            assert found_money

        finally:
            tmp_path.unlink()

    def test_export_without_cell_descriptions(self):
        """Verify include_cell_descriptions=False excludes sheet 2."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service,
                include_cell_descriptions=False
            )

            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames

            assert len(sheet_names) == 2
            assert "Matrix View" in sheet_names
            assert "Cell Descriptions" not in sheet_names
            assert "Delivery Details" in sheet_names

        finally:
            tmp_path.unlink()

    def test_export_without_delivery_details(self):
        """Verify include_delivery_details=False excludes sheet 3."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service,
                include_delivery_details=False
            )

            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames

            assert len(sheet_names) == 2
            assert "Matrix View" in sheet_names
            assert "Cell Descriptions" in sheet_names
            assert "Delivery Details" not in sheet_names

        finally:
            tmp_path.unlink()

    def test_export_with_only_matrix_view(self):
        """Verify both optional sheets can be excluded."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service,
                include_cell_descriptions=False,
                include_delivery_details=False
            )

            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames

            assert len(sheet_names) == 1
            assert "Matrix View" in sheet_names

        finally:
            tmp_path.unlink()

    def test_export_with_empty_matrix(self):
        """Verify export works with matrix containing no deliveries."""
        # Create empty matrix
        empty_matrix = self.service.create_delivery_matrix(
            components=[self.legislature.id, self.dept_ed.id],
            description="Empty test matrix",
            label="Empty Matrix"
        )

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                empty_matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)

            # Matrix View should exist with headers
            ws_matrix = wb["Matrix View"]
            assert ws_matrix.cell(1, 1).value == "Source \\ Target"
            assert ws_matrix.cell(1, 2).value == "Legislature"
            assert ws_matrix.cell(1, 3).value == "Department of Education"

            # Cell Descriptions should have only header row
            ws_desc = wb["Cell Descriptions"]
            assert ws_desc.max_row == 1

            # Delivery Details should have only header row
            ws_detail = wb["Delivery Details"]
            assert ws_detail.max_row == 1

        finally:
            tmp_path.unlink()

    def test_column_widths_and_formatting(self):
        """Verify column widths are set correctly."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = Path(tmp.name)

        try:
            export_delivery_matrix_to_xlsx(
                self.matrix,
                tmp_path,
                self.service
            )

            wb = load_workbook(tmp_path)

            # Matrix View columns
            ws_matrix = wb["Matrix View"]
            assert ws_matrix.column_dimensions['A'].width == 25
            assert ws_matrix.column_dimensions['B'].width == 30

            # Cell Descriptions columns
            ws_desc = wb["Cell Descriptions"]
            assert ws_desc.column_dimensions['A'].width == 25
            assert ws_desc.column_dimensions['B'].width == 25
            assert ws_desc.column_dimensions['C'].width == 60
            assert ws_desc.column_dimensions['D'].width == 15

            # Delivery Details columns
            ws_detail = wb["Delivery Details"]
            assert ws_detail.column_dimensions['A'].width == 25
            assert ws_detail.column_dimensions['B'].width == 25
            assert ws_detail.column_dimensions['C'].width == 15
            assert ws_detail.column_dimensions['D'].width == 50

        finally:
            tmp_path.unlink()
