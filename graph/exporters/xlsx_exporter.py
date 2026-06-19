"""
XLSX Exporter for Hayden-compliant Delivery Matrices.

Exports Social Fabric Matrix delivery matrices to Excel format matching
Hayden's published format (Hayden 2013 Koch/TD Ameritrade paper).

Three-sheet format:
1. Matrix View - N×N square matrix with component labels
2. Cell Descriptions - Narrative deliverables for each non-empty cell
3. Delivery Details - Tabular breakdown of all deliveries
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_delivery_matrix_to_xlsx(
    matrix: Any,  # SFMDeliveryMatrix type
    filepath: Path,
    service: Any,  # SFMService type
    include_cell_descriptions: bool = True,
    include_delivery_details: bool = True
) -> None:
    """
    Export square N×N delivery matrix to Excel per Hayden's published format.

    Creates a three-sheet workbook:
    - Sheet 1: Matrix View (N×N with component labels, color-coded)
    - Sheet 2: Cell Descriptions (narrative deliverables)
    - Sheet 3: Delivery Details (tabular data)

    Args:
        matrix: SFMDeliveryMatrix instance to export
        filepath: Path to save .xlsx file
        service: SFMService instance for reading component nodes
        include_cell_descriptions: Include Sheet 2 (default: True)
        include_delivery_details: Include Sheet 3 (default: True)

    Example:
        >>> from pathlib import Path
        >>> export_delivery_matrix_to_xlsx(
        ...     matrix=matrix,
        ...     filepath=Path("nebraska_k12_finance.xlsx"),
        ...     service=service
        ... )
    """
    # Build component label list
    labels = []
    for comp_id in matrix.components:
        node = service.repository.read_node(comp_id)
        labels.append(node.label if node else str(comp_id))

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Matrix View
    ws_matrix = wb.create_sheet("Matrix View", 0)
    _write_matrix_view(ws_matrix, matrix, labels, service)

    # Sheet 2: Cell Descriptions (Hayden deliverable)
    if include_cell_descriptions:
        ws_desc = wb.create_sheet("Cell Descriptions", 1)
        _write_cell_descriptions(ws_desc, matrix, service)

    # Sheet 3: Delivery Details
    if include_delivery_details:
        ws_detail = wb.create_sheet("Delivery Details", 2)
        _write_delivery_details(ws_detail, matrix, service)

    # Save workbook
    wb.save(filepath)


def _write_matrix_view(ws, matrix, labels, service):
    """
    Write Sheet 1: Matrix View.

    Format:
    - Row 1: Column headers (component labels)
    - Column A: Row headers (component labels)
    - Cells: Concatenated delivery descriptions
    - Color-coding by dominant delivery type
    """
    # Define colors for delivery types
    DELIVERY_COLORS = {
        "money": "C6EFCE",      # Light green
        "rule": "FFC7CE",       # Light red
        "authority": "FFEB9C",  # Light yellow
        "energy": "D9E1F2",     # Light blue
        "pollution": "F4B084",  # Light orange
        "information": "E2EFDA", # Light mint
        "default": "F2F2F2"     # Light gray
    }

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Border styling
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write column headers (row 1)
    ws.cell(1, 1).value = "Source \\ Target"
    ws.cell(1, 1).fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = header_alignment
    ws.cell(1, 1).border = thin_border

    for col_idx, label in enumerate(labels, start=2):
        cell = ws.cell(1, col_idx)
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write row headers and matrix data
    for row_idx, src_id in enumerate(matrix.components, start=2):
        # Row header (column A)
        src_label = labels[row_idx - 2]
        cell = ws.cell(row_idx, 1)
        cell.value = src_label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

        # Matrix cells
        for col_idx, tgt_id in enumerate(matrix.components, start=2):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Get cell data
            matrix_cell = matrix.get_cell(src_id, tgt_id)

            if matrix_cell and matrix_cell.deliveries:
                # Concatenate delivery summaries
                delivery_summaries = []
                delivery_types = []

                for d in matrix_cell.deliveries:
                    summary = f"{d.delivery_type}: {d.delivery_content[:60]}"
                    if len(d.delivery_content) > 60:
                        summary += "..."
                    delivery_summaries.append(summary)
                    delivery_types.append(d.delivery_type)

                cell.value = "\n".join(delivery_summaries)

                # Color-code by dominant delivery type
                if delivery_types:
                    dominant_type = max(set(delivery_types), key=delivery_types.count)
                    color = DELIVERY_COLORS.get(dominant_type, DELIVERY_COLORS["default"])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            else:
                cell.value = ""

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    for col_idx in range(2, len(labels) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 30

    # Adjust row heights
    for row_idx in range(2, len(labels) + 2):
        ws.row_dimensions[row_idx].height = 60


def _write_cell_descriptions(ws, matrix, service):
    """
    Write Sheet 2: Cell Descriptions.

    Hayden methodology treats cell descriptions as canonical deliverables,
    not optional metadata. Each non-empty cell gets full narrative description.

    Format:
    - Source | Target | Description | Delivery Count
    """
    # Header row
    headers = ["Source", "Target", "Description", "Delivery Count"]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_idx = 2
    for (src_id, tgt_id), cell_obj in sorted(matrix.cells.items()):
        if cell_obj.deliveries:
            # Get component labels
            src_node = service.repository.read_node(src_id)
            tgt_node = service.repository.read_node(tgt_id)
            src_label = src_node.label if src_node else str(src_id)
            tgt_label = tgt_node.label if tgt_node else str(tgt_id)

            # Write row
            ws.cell(row_idx, 1).value = src_label
            ws.cell(row_idx, 2).value = tgt_label
            ws.cell(row_idx, 3).value = cell_obj.cell_description
            ws.cell(row_idx, 4).value = len(cell_obj.deliveries)

            # Apply borders and alignment
            for col in range(1, 5):
                cell = ws.cell(row_idx, col)
                cell.border = thin_border
                if col == 3:  # Description column
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif col == 4:  # Count column
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15


def _write_delivery_details(ws, matrix, service):
    """
    Write Sheet 3: Delivery Details.

    Comprehensive tabular breakdown of all deliveries with metadata.

    Format:
    - Source | Target | Delivery Type | Content | Quantity | Units | Temporal Rate | Threshold
    """
    # Header row
    headers = [
        "Source", "Target", "Delivery Type", "Content",
        "Quantity", "Units", "Temporal Rate", "Threshold", "Certainty"
    ]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row_idx = 2
    for (src_id, tgt_id), cell_obj in sorted(matrix.cells.items()):
        # Get component labels
        src_node = service.repository.read_node(src_id)
        tgt_node = service.repository.read_node(tgt_id)
        src_label = src_node.label if src_node else str(src_id)
        tgt_label = tgt_node.label if tgt_node else str(tgt_id)

        for delivery in cell_obj.deliveries:
            # Write row
            ws.cell(row_idx, 1).value = src_label
            ws.cell(row_idx, 2).value = tgt_label
            ws.cell(row_idx, 3).value = delivery.delivery_type
            ws.cell(row_idx, 4).value = delivery.delivery_content
            ws.cell(row_idx, 5).value = delivery.quantity
            ws.cell(row_idx, 6).value = delivery.units
            ws.cell(row_idx, 7).value = delivery.temporal_rate
            ws.cell(row_idx, 8).value = delivery.threshold
            ws.cell(row_idx, 9).value = delivery.certainty

            # Apply borders and alignment
            for col in range(1, 10):
                cell = ws.cell(row_idx, col)
                cell.border = thin_border
                if col == 4:  # Content column
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif col in [5, 8, 9]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
